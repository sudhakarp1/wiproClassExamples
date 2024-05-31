
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Players")

wb.active = wb['Players']

ws = wb.active

cell = ws["A5"]

data = [
    ['PlayerName', 'Age', 'Runs', 'Matches', 'Country'],
    ['Sachin', 49, 38650, 950, 'India'],
    ['Lara', 51, 24700, 685, 'West Indies'],
    ['Ponting', 46, 32800, 705, 'Australia' ],
    ['Jayasurya', 48, 26200, 587, 'Srilanka'],
    ['Inzamam', 47, 29450, 689,'Pakistan'],
]

for row in data:
    ws.append(row)

wb.save("FirstExcel.xlsx")

