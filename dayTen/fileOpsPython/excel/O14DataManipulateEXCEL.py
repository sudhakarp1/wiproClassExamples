
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

ws = wb.active

for row in ws.iter_rows(min_row=7, min_col=1, max_col=1, max_row=11):
    for cell in row:
        if cell.value == "Ponting":
            cell.value = "ricky ponting".upper()

wb.save("FirstExcel.xlsx")
