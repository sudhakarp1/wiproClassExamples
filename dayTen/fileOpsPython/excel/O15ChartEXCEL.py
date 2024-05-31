
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart3D
from openpyxl.chart.label import DataLabelList

wb = load_workbook("FirstExcel.xlsx")
wb.create_sheet("Chart_Example")

wb.active= wb['Chart_Example']
ws = wb.active

data = [
    ('Products', 'Sales'),
    ('Pepsi', 450),
    ('Coke', 320),
    ('Sprite', 250),
    ('Mirinda', 300),
    ('Thumbs up', 475),
    ('Fanta', 230),
    ('Slice', 285)
]

for row in data:
    ws.append(row)

dataref = Reference(ws, min_row=2, min_col=2, max_row=8)
labelref = Reference(ws, min_row=2, min_col=2, max_row=8)

chart = BarChart3D()
chart.add_data(dataref)
chart.set_categories(labelref)
chart.x_axis.title = "Products"
chart.y_axis.title = "Sales in Lakhs"
chart.title = "Baverage Sales"
chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True

ws.add_chart(chart, "E9")
wb.save("FirstExcel.xlsx")
