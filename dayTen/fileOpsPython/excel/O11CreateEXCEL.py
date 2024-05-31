
from openpyxl import Workbook
from datetime import datetime
wb = Workbook()

ws = wb.active

ws.title = "MyFirstWorkSheet"

ws["C5"] = "Hello World"
ws["C7"] = 85650
ws["C9"] = datetime.now()

wb.save("FirstExcel.xlsx")
